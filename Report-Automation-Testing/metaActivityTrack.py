import requests
import json
import pandas as pd
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv
from openai import AzureOpenAI
import psycopg2
from psycopg2 import pool
import hashlib
from pytz import timezone
from global_config import get_temp_dir

load_dotenv()

AZURE_OPENAI_API_KEY = os.getenv('AZURE_OPENAI_API_KEY')
AZURE_OPENAI_ENDPOINT = os.getenv('AZURE_OPENAI_ENDPOINT')
AZURE_DEPLOYMENT = os.getenv('AZURE_DEPLOYMENT')
AZURE_OPENAI_API_VERSION = os.getenv('AZURE_OPENAI_API_VERSION', '2024-12-01-preview')

_azure_openai_client = None

def get_azure_openai_client():
    global _azure_openai_client
    if _azure_openai_client is None:
        _azure_openai_client = AzureOpenAI(
            api_key=AZURE_OPENAI_API_KEY,
            api_version=AZURE_OPENAI_API_VERSION,
            azure_endpoint=AZURE_OPENAI_ENDPOINT,
        )
    return _azure_openai_client

SUMMARY_PROMPT = (
    "You are a Performance Marketing Manager reviewing Meta Ads activity logs. "
    "Summarize the following change log in a short, clear, and consistent manner, focusing only on the essential details. "
    "For ad creative changes, specify only the fields that were modified—do not include the entire creative or unnecessary details. "
    "Avoid options or alternatives; provide a single, direct summary suitable for a daily report. "
    "Maintain a professional and concise tone.\n\n"
    "Change log:\n{change_text}"
)

def llm_summarize(change_text):
    cached = get_cached_summary(change_text)
    if cached:
        print(f"[Azure OpenAI Summarization] Cache hit for: {change_text}")
        return cached
    try:
        print(f"[Azure OpenAI Summarization] Input: {change_text}")
        client = get_azure_openai_client()
        response = client.chat.completions.create(
            model=AZURE_DEPLOYMENT,
            messages=[{'role': 'user', 'content': SUMMARY_PROMPT.format(change_text=change_text)}],
            max_completion_tokens=500,
        )
        summary = (response.choices[0].message.content or '').strip()
        print(f"[Azure OpenAI Summarization] Output: {summary}")
        set_cached_summary(change_text, summary)
        return summary
    except Exception as e:
        print(f"[Azure OpenAI Summarization] Error: {e}")
        return f"Azure OpenAI error: {e}"

# --- CONFIGURATION ---
ACCESS_TOKEN = 'EAAJ1b4rDAIQBO4WmTAhBdhLyTCZCfBVvOr2iAbYMLVWpIKfZAP2bWVp49ucZBzlRTpud5ZCRZB7Ae6pcEleUrZCgfER9enRiPQdCFcGW6TRHXtfcfNO5IVFBZCOj1HDZBaIp8EFJL2ZAZCABXYVrV5Qg7lWZAjuVuMfFX1uZBfPUPATvJ0OwY5w7tcoXqmK6LNwZBtQZDZD'
ACT_ID = '1685189008684458'

# Calculate yesterday and today dates in UTC for API calls
yesterday_utc = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
today_utc = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
print(yesterday_utc , today_utc)

SINCE = yesterday_utc.strftime('%Y-%m-%dT%H:%M:%S')
UNTIL = today_utc.replace(hour=23, minute=59, second=59).strftime('%Y-%m-%dT%H:%M:%S')
print(SINCE , UNTIL)

FIELDS = 'object_id,object_name,object_type,event_type,changed_fields,extra_data,actor_id,actor_name,event_time'
FILTERING = '[{"field":"event_type","operator":"IN","value":["UPDATE","CREATE"]}]'
LIMIT = 3000

# --- seleric_db DB CONFIGURATION ---
PG_HOST = os.getenv('DB_HOST', '72.61.228.168')
PG_PORT = os.getenv('DB_PORT', '5432')
PG_USER = os.getenv('DB_USER', 'admin_seleric')
PG_PASSWORD = os.getenv('DB_PASSWORD', 'SelericDB246')
PG_DATABASE = os.getenv('DB_NAME', 'seleric_db')

# Global connection pool
_connection_pool = None

def get_pg_conn():
    global _connection_pool
    if _connection_pool is None:
        _connection_pool = pool.SimpleConnectionPool(
            minconn=1,
            maxconn=10,
            host=PG_HOST,
            port=PG_PORT,
            user=PG_USER,
            password=PG_PASSWORD,
            dbname=PG_DATABASE
        )
    return _connection_pool.getconn()

def return_pg_conn(conn):
    global _connection_pool
    if _connection_pool:
        _connection_pool.putconn(conn)

def init_pg_cache():
    conn = get_pg_conn()
    try:
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS gemini_summary_cache (
                hash TEXT PRIMARY KEY,
                change_text TEXT,
                summary TEXT
            )
        ''')
        conn.commit()
        cur.close()
    finally:
        return_pg_conn(conn)
init_pg_cache()

def get_cache_key(change_text):
    return hashlib.sha256(change_text.encode('utf-8')).hexdigest()

def get_cached_summary(change_text):
    key = get_cache_key(change_text)
    conn = get_pg_conn()
    try:
        cur = conn.cursor()
        cur.execute('SELECT summary FROM gemini_summary_cache WHERE hash=%s', (key,))
        row = cur.fetchone()
        cur.close()
        return row[0] if row else None
    finally:
        return_pg_conn(conn)

def set_cached_summary(change_text, summary):
    key = get_cache_key(change_text)
    conn = get_pg_conn()
    try:
        cur = conn.cursor()
        cur.execute('INSERT INTO gemini_summary_cache (hash, change_text, summary) VALUES (%s, %s, %s) ON CONFLICT (hash) DO UPDATE SET summary=EXCLUDED.summary', (key, change_text, summary))
        conn.commit()
        cur.close()
    finally:
        return_pg_conn(conn)

# --- API REQUEST (moved to function level to avoid import-time errors) ---
def fetch_activity_data():
    """Fetch activity data from Meta API"""
    url = f'https://graph.facebook.com/v19.0/act_{ACT_ID}/activities'
    params = {
        'access_token': ACCESS_TOKEN,
        'fields': FIELDS,
        'since': SINCE,
        'until': UNTIL,
        'filtering': FILTERING,
        'limit': LIMIT
    }
    
    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        print(f"Error fetching activity data: {e}")
        return {}

# --- PARSE DATA ---
def summarize_change(event_type, extra):
    """
    Create a human-readable summary of the change based on event_type and extra_data.
    """
    if not extra:
        return "No extra_data"
    # Try to parse extra_data
    try:
        if isinstance(extra, str):
            extra = json.loads(extra)
    except Exception:
        return f"Unparsable extra_data: {extra}"

    # Campaign status change
    if event_type == 'update_campaign_run_status':
        old = extra.get('old_value', '')
        new = extra.get('new_value', '')
        return f"Campaign status changed: {old} → {new}"
    # Ad creative update
    elif event_type == 'update_ad_creative':
        old = extra.get('old_value', [])
        new = extra.get('new_value', [])
        # Try to extract main creative fields
        def join_vals(vals):
            if isinstance(vals, list):
                return ' | '.join(str(x) for x in vals if x)
            return str(vals)
        return f"Ad creative updated: [{join_vals(old)}] → [{join_vals(new)}]"
    # General case: show what fields changed
    elif 'old_value' in extra and 'new_value' in extra:
        return f"Changed: {extra['old_value']} → {extra['new_value']}"
    # Fallback: dump the extra_data
    return json.dumps(extra, ensure_ascii=False)

def convert_utc_to_ist(utc_time_str):
    """
    Convert UTC time string to IST timezone
    """
    try:
        # Parse UTC time
        utc_time = datetime.strptime(utc_time_str, '%Y-%m-%dT%H:%M:%S%z')
        # Convert to IST
        ist_tz = timezone('Asia/Kolkata')
        ist_time = utc_time.astimezone(ist_tz)
        return ist_time
    except Exception as e:
        print(f"Error converting time {utc_time_str}: {e}")
        return None

def is_today_ist(ist_time):
    """
    Check if the IST time is today
    """
    if ist_time is None:
        return False
    
    today_ist = datetime.now(timezone('Asia/Kolkata')).date()
    return ist_time.date() == today_ist

# Fetch data at module level with error handling
try:
    data = fetch_activity_data()
except Exception as e:
    print(f"Error fetching activity data at module level: {e}")
    data = {}

rows = []
for item in data.get('data', []):
    # Convert UTC event_time to IST
    utc_event_time = item.get('event_time')
    ist_event_time = convert_utc_to_ist(utc_event_time)
    
    # Only include events that happened today in IST
    if ist_event_time and is_today_ist(ist_event_time):
        base = {
            'object_id': item.get('object_id'),
            'object_name': item.get('object_name'),
            'object_type': item.get('object_type'),
            'event_type': item.get('event_type'),
            'actor_id': item.get('actor_id'),
            'actor_name': item.get('actor_name'),
            'event_time_utc': utc_event_time,
            'event_time_ist': ist_event_time.strftime('%Y-%m-%d %H:%M:%S %Z') if ist_event_time else None,
        }
        extra_data = item.get('extra_data')
        summary = summarize_change(item.get('event_type'), extra_data)
        base['change_summary'] = summary
        rows.append(base)

def generate_meta_activity_excel():
    """
    Generates the meta activity Excel file and returns the file path.
    Returns:
        str: Path to the generated Excel file, or None if no data.
    """
    df = pd.DataFrame(rows)
    if not df.empty:
        # Convert IST time string back to datetime for sorting
        df['event_time_ist_dt'] = pd.to_datetime(
            df['event_time_ist'].str.replace(' IST', ''), errors='coerce'
        ).dt.tz_localize('Asia/Kolkata')

        # Find the latest event_time for each object_id
        latest_event_times = df.groupby('object_id')['event_time_ist_dt'].max().reset_index()
        latest_event_times = latest_event_times.rename(columns={'event_time_ist_dt': 'latest_event_time'})
        df = df.merge(latest_event_times, on='object_id', how='left')

        # Sort by latest_event_time (descending), then by object_id, then by event_time (descending within group)
        df = df.sort_values(['latest_event_time', 'object_id', 'event_time_ist_dt'], ascending=[False, True, False])
        df = df.drop(columns=['latest_event_time', 'event_time_ist_dt'])

        # For each group, keep object_id, object_name, object_type only in the first row
        for col in ['object_id', 'object_name', 'object_type']:
            df[col] = df.groupby('object_id')[col].transform(lambda x: [x.iloc[0]] + [''] * (len(x) - 1))

        # Remove actor_id column if present
        if 'actor_id' in df.columns:
            df = df.drop(columns=['actor_id'])

        # Add LLM-generated summary column using Azure OpenAI with optimized caching
        print("Processing LLM summaries with optimized caching...")
        df['llm_summary'] = df['change_summary'].apply(llm_summarize)

        # Map object_type values for readability BEFORE exporting
        object_type_map = {
            'CAMPAIGN_GROUP': 'Campaign',
            'CAMPAIGN': 'Adset',
            'ADGROUP': 'Ad',
        }
        df['object_type'] = df['object_type'].replace(object_type_map)

        # Reorder columns for readability
        columns = [
            'object_name', 'object_type', 'event_type',
            'event_time_ist', 'actor_name', 'llm_summary'
        ]
        df = df[columns]

        # Export to Excel
        _temp_dir = get_temp_dir()
        excel_path = os.path.join(_temp_dir, 'activity_history.xlsx')
        df.to_excel(excel_path, index=False)
        print(f"Exported today's changes (IST) to {excel_path}")

        # --- Format Excel: auto-fit columns and bold latest change in each group ---
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Bold the latest change in every object_name group
        from collections import defaultdict
        latest_row_by_object = defaultdict(int)
        object_name_col = 1  # Assuming 'object_name' is the first column
        event_time_col = 4   # Assuming 'event_time_ist' is the 4th column

        # Find the latest row for each object_name
        for row in range(2, ws.max_row + 1):
            object_name = ws.cell(row=row, column=object_name_col).value
            event_time = ws.cell(row=row, column=event_time_col).value
            if object_name:
                if (object_name not in latest_row_by_object or
                    (event_time and ws.cell(row=latest_row_by_object[object_name], column=event_time_col).value < event_time)):
                    latest_row_by_object[object_name] = row

        # Apply bold font to the latest row in each group
        for row in latest_row_by_object.values():
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).font = Font(bold=True)

        wb.save(excel_path)
        print("Formatted Excel: auto-fit columns and bolded latest change in each group.")

        # --- Merge cells for object_id, object_name, object_type in Excel ---
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        current_id = None
        start_row = 2  # 1-based index, row 1 is header
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1), start=2):
            cell_value = row[0].value
            if cell_value != current_id:
                # If not the first group, merge previous group
                if current_id is not None and start_row < i - 1:
                    for col in range(1, 4):  # Columns A, B, C (1-based)
                        ws.merge_cells(start_row=start_row, start_column=col, end_row=i-1, end_column=col)
                current_id = cell_value
                start_row = i
        # Merge the last group
        if current_id is not None and start_row < ws.max_row:
            for col in range(1, 4):
                ws.merge_cells(start_row=start_row, start_column=col, end_row=ws.max_row, end_column=col)

        wb.save(excel_path)
        print("Merged cells for object_id, object_name, and object_type in Excel output.")
        return excel_path
    else:
        print("No data to export for today (IST).")
        return None

def cleanup_connections():
    """Clean up database connections"""
    global _connection_pool
    if _connection_pool:
        _connection_pool.closeall()
        _connection_pool = None

if __name__ == "__main__":
    try:
        generate_meta_activity_excel()
    finally:
        cleanup_connections()

# Note: You need to install the required packages with: pip install openai pytz
